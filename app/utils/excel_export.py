"""Shared .xlsx builder for the cross-portal "Export Activity Report" feature. Every portal
(Super Admin, Vendor, Warehouse, Retail) exports through this ONE function so the visual
template -- logo, summary cards, entity-details card, activity table -- stays identical across
portals; only the portal label, entity details, and rows differ per caller.

Presentation only: card-style corporate dashboard layout. No business/query logic lives here --
callers (the 4 export endpoints) still decide what portal_label/details/period_label/rows are;
this module only decides how to draw them."""
import io
from datetime import datetime
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.models.audit import AuditLog

_LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "brfid-logo.png"
_BUSINESS_LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "brfid-logo-new.jpeg"

# ---- palette: restrained corporate set (navy primary, light-blue/gray secondary, a single
# purple/teal accent) -- reused everywhere so no card/section improvises its own colors. ----
_NAVY = "1F3864"
_NAVY_TEXT = "1F3864"
_GRAY_TEXT = "404040"
_MUTED_TEXT = "7F7F7F"
_LIGHT_GRID = "E3E7ED"
_CARD_FILLS = ["DCE6F1", "E8E3F2", "D8EEF0", "F2F2F2"]  # light blue / light purple / light teal / light gray
_CARD_ACCENTS = ["2E5C8A", "6B4E9C", "1E8A82", "595959"]
_DETAILS_FILL = "F7F9FC"
_TABLE_HEADER_FILL = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_TABLE_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10.5)
_ZEBRA_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
_THIN_GRID = Side(style="thin", color=_LIGHT_GRID)
_THIN_BORDER = Border(left=_THIN_GRID, right=_THIN_GRID, top=_THIN_GRID, bottom=_THIN_GRID)

_COL_WIDTHS = {"A": 20, "B": 24, "C": 52, "D": 24}
_TABLE_HEADERS = ["Date/Time", "Action Type", "Description", "Performed By"]
_TABLE_COL_COUNT = len(_TABLE_HEADERS)
_CARD_ICONS = ["\U0001F4C5", "\U0001F551", "\U0001F3F7", "\U0001F4CA"]  # calendar / clock / tag / bar-chart
_LOCATION_KEYS = {"address", "city", "state", "pincode", "postal code", "zip"}
_DESC_CHARS_PER_LINE = 48  # rough fit for column C's width at 10.5pt -- drives wrapped row height
_ROW_LINE_HEIGHT = 14
_MAX_ROW_HEIGHT = 90


def _humanize_action(action_type: str) -> str:
    """Some action_type values already read as a phrase ("Vendor Selected"); older ones are
    SCREAMING_SNAKE_CASE ("VENDOR_APPROVED") -- normalize the latter so the table doesn't mix
    two different casing styles in the same column."""
    if "_" in action_type or action_type.isupper():
        return action_type.replace("_", " ").title()
    return action_type


def _box_border(ws, first_row: int, last_row: int, first_col: int, last_col: int, color: str = _LIGHT_GRID) -> None:
    """Draws a rectangular outline (a "card") around a cell range without merging it."""
    side = Side(style="thin", color=color)
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            top = side if r == first_row else None
            bottom = side if r == last_row else None
            left = side if c == first_col else None
            right = side if c == last_col else None
            if top or bottom or left or right:
                cell.border = Border(top=top or cell.border.top, bottom=bottom or cell.border.bottom,
                                      left=left or cell.border.left, right=right or cell.border.right)


def _section_title(ws, row: int, text: str, span_cols: int = 4) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row=row, column=1, value=text.upper())
    cell.font = Font(bold=True, size=12, color=_NAVY_TEXT)
    cell.alignment = Alignment(vertical="center")
    for col in range(1, span_cols + 1):
        ws.cell(row=row, column=col).border = Border(bottom=Side(style="medium", color=_NAVY))
    ws.row_dimensions[row].height = 22
    return row + 1


def _summary_cards(ws, row: int, cards: list[tuple[str, str]]) -> int:
    """cards: list of (label, value) pairs, one per column A-D."""
    label_row, value_row = row, row + 1
    ws.row_dimensions[label_row].height = 16
    ws.row_dimensions[value_row].height = 28
    for idx, (label, value) in enumerate(cards):
        col = idx + 1
        fill = PatternFill(start_color=_CARD_FILLS[idx], end_color=_CARD_FILLS[idx], fill_type="solid")
        accent = _CARD_ACCENTS[idx]
        label_cell = ws.cell(row=label_row, column=col, value=f"{_CARD_ICONS[idx]}  {label}")
        label_cell.font = Font(bold=True, size=8.5, color=accent)
        label_cell.fill = fill
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = ws.cell(row=value_row, column=col, value=value)
        value_cell.font = Font(bold=True, size=15, color=_NAVY_TEXT)
        value_cell.fill = fill
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _box_border(ws, label_row, value_row, col, col, color=accent)
    return value_row + 2


def _details_card(ws, row: int, portal_label: str, details: dict[str, str]) -> int:
    row = _section_title(row=row, ws=ws, text=f"{portal_label} Details")
    card_start = row

    left_items = [(k, v) for k, v in details.items() if k.lower() not in _LOCATION_KEYS]
    right_items = [(k, v) for k, v in details.items() if k.lower() in _LOCATION_KEYS]

    for left, right in zip_longest(left_items, right_items, fillvalue=(None, None)):
        is_address_row = (left[0] or "").lower() == "address" or (right[0] or "").lower() == "address"
        if left[0] is not None:
            lk = ws.cell(row=row, column=1, value=left[0])
            lk.font = Font(bold=True, size=9.5, color=_GRAY_TEXT)
            lk.alignment = Alignment(vertical="center", indent=1)
            lv = ws.cell(row=row, column=2, value=left[1] if left[1] not in (None, "") else "-")
            lv.font = Font(size=10, color=_NAVY_TEXT)
            lv.alignment = Alignment(vertical="center", wrap_text=is_address_row, indent=1)
        if right[0] is not None:
            rk = ws.cell(row=row, column=3, value=right[0])
            rk.font = Font(bold=True, size=9.5, color=_GRAY_TEXT)
            rk.alignment = Alignment(vertical="center", indent=1)
            rv = ws.cell(row=row, column=4, value=right[1] if right[1] not in (None, "") else "-")
            rv.font = Font(size=10, color=_NAVY_TEXT)
            rv.alignment = Alignment(vertical="center", wrap_text=is_address_row, indent=1)
        ws.row_dimensions[row].height = 34 if is_address_row else 18
        row += 1

    card_end = row - 1
    fill = PatternFill(start_color=_DETAILS_FILL, end_color=_DETAILS_FILL, fill_type="solid")
    for r in range(card_start, card_end + 1):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill
    _box_border(ws, card_start, card_end, 1, 4)
    # Left accent stripe, a common dashboard-card touch -- kept to one thin colored edge only.
    for r in range(card_start, card_end + 1):
        cell = ws.cell(row=r, column=1)
        cell.border = Border(left=Side(style="thick", color=_NAVY), top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right)
    return card_end + 2


def _build_activity_sheet(wb, portal_label: str, period_label: str, rows: list[AuditLog]) -> None:
    """Activity History gets its OWN tab, deliberately -- keeping it off the Summary sheet
    means freeze_panes only ever has to pin this sheet's own small 4-row header, not the
    Summary sheet's much taller logo+cards+details block. That block is tall enough that
    freezing beneath it (the previous single-sheet design) could pin most or all of a normal
    screen's height, leaving little or nothing to actually scroll -- this tab is always fully
    scrollable regardless of how tall the Summary sheet's header is."""
    ws = wb.create_sheet("Activity History")
    ws.sheet_view.showGridLines = False
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    title_cell = ws.cell(row=1, column=1, value=f"{portal_label.upper()} ACTIVITY HISTORY")
    title_cell.font = Font(bold=True, size=14, color=_NAVY_TEXT)
    title_cell.alignment = Alignment(horizontal="center")
    sub_cell = ws.cell(row=2, column=1, value=f"{period_label}  |  {len(rows)} activit{'y' if len(rows) == 1 else 'ies'}")
    sub_cell.font = Font(italic=True, size=10, color=_MUTED_TEXT)
    sub_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_TABLE_COL_COUNT)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_TABLE_COL_COUNT)

    header_row = 4
    for col_idx, header in enumerate(_TABLE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _TABLE_HEADER_FONT
        cell.fill = _TABLE_HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[header_row].height = 22
    row_cursor = header_row + 1

    if not rows:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=_TABLE_COL_COUNT)
        cell = ws.cell(row=row_cursor, column=1, value="No activity found for the selected period.")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(italic=True, color=_MUTED_TEXT)
        ws.row_dimensions[row_cursor].height = 24
        for col_idx in range(1, _TABLE_COL_COUNT + 1):
            ws.cell(row=row_cursor, column=col_idx).border = _THIN_BORDER
        row_cursor += 1
    else:
        for i, row in enumerate(rows):
            description = row.description or "-"
            values = [
                row.occurred_at.strftime("%d %b %Y  %H:%M"),
                _humanize_action(row.action_type),
                description,
                getattr(row, "performed_by_name", "-"),
            ]
            line_count = max(1, -(-len(description) // _DESC_CHARS_PER_LINE))  # ceil division
            ws.row_dimensions[row_cursor].height = min(_MAX_ROW_HEIGHT, max(28, _ROW_LINE_HEIGHT * line_count + 12))
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.font = Font(size=10.5, color=_GRAY_TEXT)
                if i % 2 == 1:
                    cell.fill = _ZEBRA_FILL
                if col_idx == 3:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")
            row_cursor += 1
        ws.auto_filter.ref = f"A{header_row}:{chr(64 + _TABLE_COL_COUNT)}{row_cursor - 1}"

    # Freeze ONLY this sheet's own 4-row header -- the rest of the window is free to scroll.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)


def build_activity_report_workbook(
    portal_label: str,
    details: dict[str, str] | None,
    period_label: str,
    rows: list[AuditLog],
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # ---- header band: big title centered, logo right ----
    ws.merge_cells("A1:B2")
    title_cell = ws.cell(row=1, column=1, value=f"{portal_label.upper()} ACTIVITY REPORT")
    title_cell.font = Font(bold=True, size=18, color=_NAVY_TEXT)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    if _LOGO_PATH.exists():
        img = XLImage(str(_LOGO_PATH))
        original_width, original_height = img.width, img.height
        img.width = 190
        img.height = int(190 * original_height / original_width) if original_width else 67
        ws.add_image(img, "D1")
    else:
        ws.cell(row=1, column=4, value="Britannia RFID").font = Font(bold=True, size=14, color=_NAVY_TEXT)

    for col in range(1, _TABLE_COL_COUNT + 1):
        ws.cell(row=4, column=col).border = Border(bottom=Side(style="medium", color=_NAVY))
    row_cursor = 6

    # ---- summary cards: report period / generated on / entity code / total activities ----
    if ":" in period_label:
        period_key, period_value = period_label.split(":", 1)
        period_key, period_value = period_key.strip().upper(), period_value.strip()
    else:
        period_key, period_value = "REPORT PERIOD", period_label

    now = datetime.now()
    code_key = next((k for k in (details or {}) if k.lower().endswith("code")), None)
    if code_key:
        code_label, code_value = f"{portal_label.upper()} CODE", details[code_key]
    else:
        code_label, code_value = "SCOPE", "All Portals"

    row_cursor = _summary_cards(ws, row_cursor, [
        (period_key, period_value),
        ("GENERATED ON", f"{now:%d %b %Y}\n{now:%H:%M}"),
        (code_label, code_value),
        ("TOTAL ACTIVITIES", len(rows)),
    ])

    # ---- entity details card (omitted for Super Admin, which has no single entity) ----
    if details:
        row_cursor = _details_card(ws, row_cursor, portal_label, details)

    note = ws.cell(row=row_cursor, column=1, value="▶  See the \"Activity History\" tab below for the full activity list.")
    note.font = Font(italic=True, size=10, color=_MUTED_TEXT)
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=_TABLE_COL_COUNT)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)

    _build_activity_sheet(wb, portal_label, period_label, rows)
    wb.active = 0  # open on the Summary tab first

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------------------------
# Business-data report builder (Vendor/Warehouse/Retail dashboard "Export Report" -- Week/Month).
# Deliberately plain and structured, per the redesign brief: no decorative icons, no colored
# summary cards, one clearly-bordered grid per data section, consistent widths/alignment, and a
# clean separation between the header/summary block and each section's own data table.
# ---------------------------------------------------------------------------------------------

_REPORT_HEADER_FILL = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_REPORT_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10.5)
_REPORT_BODY_FONT = Font(size=10, color=_GRAY_TEXT)
_REPORT_ROW_ALT_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
_REPORT_TILE_FILL = PatternFill(start_color="F2F5FA", end_color="F2F5FA", fill_type="solid")
_REPORT_CARD_FILL = PatternFill(start_color="FAFBFC", end_color="FAFBFC", fill_type="solid")
_REPORT_ACCENT_BORDER = Side(style="thin", color="D6DCE5")
_REPORT_CONTENTS_BORDER = Border(
    left=_REPORT_ACCENT_BORDER, right=_REPORT_ACCENT_BORDER, top=_REPORT_ACCENT_BORDER, bottom=_REPORT_ACCENT_BORDER,
)


def _report_header_block(
    ws, portal_label: str, title_suffix: str, *,
    show_logo: bool = True, subtitle: str = "Britannia RFID Platform", center_align: bool = False,
) -> int:
    """Title + subtitle centered, small logo right (when show_logo) -- no rule beneath -- returns
    the next free row. The title/subtitle are always centered regardless of center_align, which
    only governs the report body (meta strip, details, contents, data tables) below."""
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value=f"{portal_label.upper()} {title_suffix}".strip())
    title_cell.font = Font(bold=True, size=18, color=_NAVY_TEXT)
    title_cell.alignment = Alignment(horizontal="center", vertical="bottom")

    ws.merge_cells("A2:D2")
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(size=10, italic=True, color=_MUTED_TEXT)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="top")

    if show_logo:
        if _BUSINESS_LOGO_PATH.exists():
            img = XLImage(str(_BUSINESS_LOGO_PATH))
            original_width, original_height = img.width, img.height
            img.width = 80
            img.height = int(80 * original_height / original_width) if original_width else 28
            ws.add_image(img, "F1")
        else:
            ws.cell(row=1, column=6, value="Britannia RFID").font = Font(bold=True, size=11, color=_NAVY_TEXT)

    ws.row_dimensions[3].height = 8
    return 6


def _report_meta_strip(ws, row: int, meta: list[tuple[str, str]], *, center_align: bool = False) -> int:
    """Two light tinted tiles side by side (label above, value below) -- a compact, modern
    stand-in for the old colored icon cards, with no icons/colors of its own."""
    h_align = "center" if center_align else "left"
    label_row, value_row = row, row + 1
    ws.row_dimensions[label_row].height = 15
    ws.row_dimensions[value_row].height = 22
    spans = [(1, 2), (3, 4)]
    for (start_col, end_col), (label, value) in zip(spans, meta):
        ws.merge_cells(start_row=label_row, start_column=start_col, end_row=label_row, end_column=end_col)
        ws.merge_cells(start_row=value_row, start_column=start_col, end_row=value_row, end_column=end_col)
        label_cell = ws.cell(row=label_row, column=start_col, value=label.upper())
        label_cell.font = Font(bold=True, size=8.5, color=_MUTED_TEXT)
        label_cell.alignment = Alignment(horizontal=h_align)
        value_cell = ws.cell(row=value_row, column=start_col, value=value)
        value_cell.font = Font(bold=True, size=12, color=_NAVY_TEXT)
        value_cell.alignment = Alignment(horizontal=h_align)
        for r in (label_row, value_row):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).fill = _REPORT_TILE_FILL
        _box_border(ws, label_row, value_row, start_col, end_col, color="D6DCE5")
    return value_row + 2


def _report_details_block(ws, row: int, heading: str, details: dict[str, str], *, center_align: bool = False) -> int:
    """One key/value pair per row (label in column A, value spanning B-D) inside a
    soft-bordered box -- Vendor Code, Name, GST, Address, City, State etc. listed one by one
    rather than packed into side-by-side pairs."""
    h_align = "center" if center_align else "left"
    heading_cell = ws.cell(row=row, column=1, value=heading.upper())
    heading_cell.font = Font(bold=True, size=11, color=_NAVY_TEXT)
    heading_cell.alignment = Alignment(horizontal=h_align)
    row += 1
    card_start = row

    for key, value in details.items():
        is_address_row = key.lower() == "address"
        key_cell = ws.cell(row=row, column=1, value=key)
        key_cell.font = Font(bold=True, size=9.5, color=_GRAY_TEXT)
        key_cell.alignment = Alignment(horizontal=h_align, vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        val_cell = ws.cell(row=row, column=2, value=value if value not in (None, "") else "-")
        val_cell.font = Font(size=10, color=_NAVY_TEXT)
        val_cell.alignment = Alignment(horizontal=h_align, wrap_text=is_address_row, vertical="center", indent=1)
        ws.row_dimensions[row].height = 38 if is_address_row else 17
        row += 1

    card_end = row - 1
    for r in range(card_start, card_end + 1):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = _REPORT_CARD_FILL
    _box_border(ws, card_start, card_end, 1, 4, color="E3E7ED")
    return card_end + 2


def _report_contents_block(ws, row: int, sections: list[dict], *, center_align: bool = False) -> int:
    """A small bordered table (matching the data sheets' own header style) listing each section
    and its record count, rather than a plain text list."""
    h_align = "center" if center_align else "left"
    heading_cell = ws.cell(row=row, column=1, value="REPORT CONTENTS")
    heading_cell.font = Font(bold=True, size=11, color=_NAVY_TEXT)
    heading_cell.alignment = Alignment(horizontal=h_align)
    row += 1

    header_row = row
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=3)
    ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row, end_column=4)
    for col, label in ((1, "Sheet"), (4, "Records")):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = _REPORT_HEADER_FONT
        cell.fill = _REPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if center_align else ("left" if col == 1 else "right"), vertical="center")
    for c in range(1, 5):
        ws.cell(row=header_row, column=c).border = _REPORT_CONTENTS_BORDER
    ws.row_dimensions[header_row].height = 19
    row += 1

    for i, section in enumerate(sections):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=4)
        count = len(section["rows"])
        name_cell = ws.cell(row=row, column=1, value=section["title"])
        name_cell.font = Font(size=10, color=_NAVY_TEXT)
        name_cell.alignment = Alignment(horizontal=h_align)
        count_cell = ws.cell(row=row, column=4, value=count)
        count_cell.font = Font(size=10, color=_GRAY_TEXT)
        count_cell.alignment = Alignment(horizontal="center" if center_align else "right")
        fill = _REPORT_ROW_ALT_FILL if i % 2 == 1 else None
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = _REPORT_CONTENTS_BORDER
            if fill:
                cell.fill = fill
        row += 1
    return row + 1


def _build_report_data_sheet(wb, title: str, columns: list[str], rows: list[list], *, center_align: bool = False) -> None:
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    col_count = len(columns)

    widths = []
    for idx, col in enumerate(columns):
        longest = len(str(col))
        for r in rows[:300]:
            value = r[idx]
            if value is not None:
                longest = max(longest, len(str(value)))
        widths.append(min(42, max(13, longest + 2)))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_row = 1
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _REPORT_HEADER_FONT
        cell.fill = _REPORT_HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center") if center_align else Alignment(vertical="center")
    ws.row_dimensions[header_row].height = 20
    row_cursor = header_row + 1

    if not rows:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=col_count)
        cell = ws.cell(row=row_cursor, column=1, value="No records found for the selected period.")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(italic=True, color=_MUTED_TEXT)
        ws.row_dimensions[row_cursor].height = 24
        for col_idx in range(1, col_count + 1):
            ws.cell(row=row_cursor, column=col_idx).border = _THIN_BORDER
    else:
        for i, values in enumerate(rows):
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.font = _REPORT_BODY_FONT
                if i % 2 == 1:
                    cell.fill = _REPORT_ROW_ALT_FILL
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if center_align else ("right" if isinstance(value, (int, float)) else "left"),
                )
            row_cursor += 1
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{row_cursor - 1}"

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)


def build_business_report_workbook(
    portal_label: str,
    entity_details: dict[str, str] | None,
    period_label: str,
    sections: list[dict],
    *,
    show_logo: bool = True,
    subtitle: str = "Britannia RFID Platform",
    center_align: bool = False,
) -> io.BytesIO:
    """Generic replacement for the activity-log report, for the Vendor/Warehouse/Retail
    dashboard export dialog. `sections` is `[{"title": str, "columns": [str, ...],
    "rows": [[...], ...]}, ...]` -- each section gets its own sheet with a plain bordered grid
    (one field per column, one record per row); the Summary sheet only holds the header, period,
    entity details, and a one-line-per-section contents index -- no business logic here, callers
    decide what sections/rows exist."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28

    row_cursor = _report_header_block(
        ws, portal_label, "BUSINESS REPORT", show_logo=show_logo, subtitle=subtitle, center_align=center_align,
    )
    row_cursor = _report_meta_strip(ws, row_cursor, [
        ("Report Period", period_label),
        ("Generated On", datetime.now().strftime("%d %b %Y, %H:%M")),
    ], center_align=center_align)

    if entity_details:
        row_cursor = _report_details_block(ws, row_cursor, f"{portal_label} Details", entity_details, center_align=center_align)

    _report_contents_block(ws, row_cursor, sections, center_align=center_align)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)

    for section in sections:
        _build_report_data_sheet(wb, section["title"], section["columns"], section["rows"], center_align=center_align)

    wb.active = 0
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
